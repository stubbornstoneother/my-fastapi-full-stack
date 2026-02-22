"""Personnel management API routes."""
import io
import os
import uuid
import zipfile
from datetime import datetime, timezone
from typing import Any

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlmodel import col, func, select

from app.api.deps import CurrentUser, SessionDep
from app.models import Message
from app.models_system import (
    BatchUpdateRequest,
    PersonInfo,
    PersonInfoCreate,
    PersonInfoPublic,
    PersonInfoUpdate,
    PersonRlistItem,
    PersonRlistRequest,
    PersonsPublic,
    Robot,
    SystemLog,
)

router = APIRouter(prefix="/system", tags=["persons"])

AVATAR_DIR = "/app/avatars"


# ===== CRUD =====

@router.get("/persons/", response_model=PersonsPublic, tags=["persons"])
def read_persons(
    session: SessionDep, current_user: CurrentUser,
    skip: int = 0, limit: int = 100,
    search: str | None = None,
    org_id: uuid.UUID | None = None,
) -> Any:
    """List personnel."""
    query = select(PersonInfo).where(PersonInfo.del_flag == "0")
    if search:
        query = query.where(
            (PersonInfo.name.contains(search))  # type: ignore
            | (PersonInfo.card_id.contains(search))  # type: ignore
        )
    if org_id:
        query = query.where(PersonInfo.sys_org_code == org_id)

    count_stmt = select(func.count()).select_from(query.subquery())
    count = session.exec(count_stmt).one()

    persons = session.exec(
        query.order_by(col(PersonInfo.created_at).desc()).offset(skip).limit(limit)
    ).all()

    return PersonsPublic(data=persons, count=count)


@router.post("/persons/", response_model=PersonInfoPublic, tags=["persons"])
def create_person(
    *, session: SessionDep, current_user: CurrentUser, person_in: PersonInfoCreate,
) -> Any:
    """Create a new person."""
    existing = session.exec(
        select(PersonInfo).where(
            PersonInfo.card_id == person_in.card_id,
            PersonInfo.del_flag == "0",
        )
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="身份证号已存在")

    person = PersonInfo.model_validate(person_in)
    session.add(person)
    session.commit()
    session.refresh(person)

    # Mark robots for user update
    _mark_robots_for_user_update(session)

    return person


@router.put("/persons/{person_id}", response_model=PersonInfoPublic, tags=["persons"])
def update_person(
    *, session: SessionDep, current_user: CurrentUser,
    person_id: uuid.UUID, person_in: PersonInfoUpdate,
) -> Any:
    """Update a person."""
    person = session.get(PersonInfo, person_id)
    if not person or person.del_flag == "1":
        raise HTTPException(status_code=404, detail="人员不存在")

    update_data = person_in.model_dump(exclude_unset=True)
    person.sqlmodel_update(update_data)
    person.updated_at = datetime.now(timezone.utc)
    session.add(person)
    session.commit()
    session.refresh(person)

    _mark_robots_for_user_update(session)

    return person


@router.delete("/persons/{person_id}", tags=["persons"])
def delete_person(
    session: SessionDep, current_user: CurrentUser, person_id: uuid.UUID,
) -> Message:
    """Soft delete a person."""
    person = session.get(PersonInfo, person_id)
    if not person:
        raise HTTPException(status_code=404, detail="人员不存在")

    person.del_flag = "1"
    person.updated_at = datetime.now(timezone.utc)
    session.add(person)
    session.commit()

    _mark_robots_for_user_update(session)

    return Message(message="删除成功")


@router.post("/persons/batch-delete", tags=["persons"])
def batch_delete_persons(
    *, session: SessionDep, current_user: CurrentUser, ids: list[uuid.UUID],
) -> Message:
    """Batch soft delete persons."""
    for pid in ids:
        person = session.get(PersonInfo, pid)
        if person:
            person.del_flag = "1"
            person.updated_at = datetime.now(timezone.utc)
            session.add(person)
    session.commit()
    _mark_robots_for_user_update(session)
    return Message(message="批量删除成功")


# ===== Batch Update =====

@router.post("/persons/batch-update", tags=["persons"])
def batch_update_persons(
    *, session: SessionDep, current_user: CurrentUser, req: BatchUpdateRequest,
) -> Message:
    """Batch update person fields (category, difficulty, age, org)."""
    for pid in req.ids:
        person = session.get(PersonInfo, pid)
        if person and person.del_flag == "0":
            if req.category is not None:
                person.category = req.category
            if req.category_name is not None:
                person.category_name = req.category_name
            if req.difficulty is not None:
                person.difficulty = req.difficulty
            if req.age is not None:
                person.age = req.age
            if req.sys_org_code is not None:
                person.sys_org_code = req.sys_org_code
            person.updated_at = datetime.now(timezone.utc)
            session.add(person)
    session.commit()
    _mark_robots_for_user_update(session)
    return Message(message="批量更新成功")


# ===== Excel Import =====

@router.post("/persons/import", tags=["persons"])
async def import_persons_excel(
    *, session: SessionDep, current_user: CurrentUser,
    file: UploadFile = File(...),
) -> Message:
    """Import persons from Excel file."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    if not ws:
        raise HTTPException(status_code=400, detail="Excel文件为空")

    headers = [cell.value for cell in ws[1]]
    field_map = {
        "姓名": "name", "性别": "gender", "年龄": "age",
        "身高": "height", "体重": "weight",
        "身份证号": "card_id", "军人证号": "soldier_id",
        "人员类别": "category_name", "类别编码": "category",
        "考核难度": "difficulty", "部职别": "title",
        "伤病": "disease",
    }

    imported = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i] in field_map:
                row_data[field_map[headers[i]]] = val

        if not row_data.get("name") or not row_data.get("card_id"):
            skipped += 1
            continue

        # Ensure age is int
        if "age" in row_data and row_data["age"] is not None:
            row_data["age"] = int(row_data["age"])
        else:
            row_data["age"] = 0

        # Set defaults
        row_data.setdefault("gender", "男")
        row_data.setdefault("soldier_id", row_data.get("card_id", ""))
        row_data.setdefault("category_name", "一类人员")
        row_data.setdefault("category", "1")
        row_data.setdefault("difficulty", "2")

        # Check if exists
        existing = session.exec(
            select(PersonInfo).where(
                PersonInfo.card_id == row_data["card_id"],
                PersonInfo.del_flag == "0",
            )
        ).first()
        if existing:
            skipped += 1
            continue

        person = PersonInfo(**row_data)
        session.add(person)
        imported += 1

    session.commit()
    if imported > 0:
        _mark_robots_for_user_update(session)

    return Message(message=f"导入完成: 成功 {imported} 条, 跳过 {skipped} 条")


@router.get("/persons/template", tags=["persons"])
def download_import_template() -> StreamingResponse:
    """Download Excel import template."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "人员信息模板"
    headers = ["姓名", "性别", "年龄", "身高", "体重", "身份证号",
               "军人证号", "人员类别", "类别编码", "考核难度", "部职别", "伤病"]
    ws.append(headers)

    # Example row
    ws.append(["张三", "男", 25, "178", "75", "130425199909121012",
               "130425199909121012", "一类人员", "1", "2", "士兵", "否"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=personnel_template.xlsx"},
    )


# ===== Avatar Upload =====

@router.post("/persons/{person_id}/avatar", tags=["persons"])
async def upload_avatar(
    *, session: SessionDep, current_user: CurrentUser,
    person_id: uuid.UUID, file: UploadFile = File(...),
) -> Message:
    """Upload avatar for a person."""
    person = session.get(PersonInfo, person_id)
    if not person or person.del_flag == "1":
        raise HTTPException(status_code=404, detail="人员不存在")

    os.makedirs(AVATAR_DIR, exist_ok=True)
    ext = os.path.splitext(file.filename or "avatar.jpg")[1] or ".jpg"
    filename = f"{person.card_id}{ext}"
    filepath = os.path.join(AVATAR_DIR, filename)

    content = await file.read()
    with open(filepath, "wb") as f:
        f.write(content)

    person.avatar_path = filepath
    person.updated_at = datetime.now(timezone.utc)
    session.add(person)
    session.commit()

    return Message(message="头像上传成功")


@router.post("/persons/avatars", tags=["persons"])
async def batch_upload_avatars(
    *, session: SessionDep, current_user: CurrentUser,
    files: list[UploadFile] = File(...),
) -> Message:
    """Batch upload avatars. Filename must be cardId."""
    os.makedirs(AVATAR_DIR, exist_ok=True)
    success = 0
    for file in files:
        if not file.filename:
            continue
        card_id = os.path.splitext(file.filename)[0]
        ext = os.path.splitext(file.filename)[1] or ".jpg"

        person = session.exec(
            select(PersonInfo).where(
                PersonInfo.card_id == card_id,
                PersonInfo.del_flag == "0",
            )
        ).first()
        if not person:
            continue

        filename = f"{card_id}{ext}"
        filepath = os.path.join(AVATAR_DIR, filename)
        content = await file.read()
        with open(filepath, "wb") as f:
            f.write(content)

        person.avatar_path = filepath
        person.updated_at = datetime.now(timezone.utc)
        session.add(person)
        success += 1

    session.commit()
    return Message(message=f"批量上传完成: 成功 {success} 条")


# ===== Robot Download APIs (供机器人调用) =====

@router.post("/userinfo/rlist", response_model=list[PersonRlistItem], tags=["persons"])
def batch_download_persons(
    *, session: SessionDep, req: PersonRlistRequest,
) -> Any:
    """Batch download person data for robots."""
    query = select(PersonInfo)

    if req.finalDate:
        try:
            final_dt = datetime.strptime(req.finalDate, "%Y-%m-%d %H:%M:%S")
            final_dt = final_dt.replace(tzinfo=timezone.utc)
            # Incremental update: return records updated after finalDate
            query = query.where(PersonInfo.updated_at > final_dt)
        except ValueError:
            pass
    else:
        # Full download: only return non-deleted records
        query = query.where(PersonInfo.del_flag == "0")

    persons = session.exec(query).all()

    # If robot code provided, update the robot's tracking
    if req.robotCode:
        robot = session.exec(
            select(Robot).where(Robot.code == req.robotCode)
        ).first()
        if robot:
            robot.user_info_version = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
            robot.needs_user_update = False
            # Clear update_user command if pending
            if robot.pending_command in ("update_user", "update_user_all"):
                robot.pending_command = None
            session.add(robot)
            session.commit()

    result = []
    for p in persons:
        result.append(PersonRlistItem(
            id=str(p.id),
            createBy=None,
            createTime=p.created_at.strftime("%Y-%m-%d %H:%M:%S") if p.created_at else None,
            updateBy=None,
            updateTime=p.updated_at.strftime("%Y-%m-%d %H:%M:%S") if p.updated_at else None,
            sysOrgCode=str(p.sys_org_code) if p.sys_org_code else None,
            name=p.name,
            gender=p.gender,
            age=p.age,
            height=p.height,
            weight=p.weight,
            cardId=p.card_id,
            soldierId=p.soldier_id,
            categoryName=p.category_name,
            difficulty=p.difficulty,
            title=p.title,
            category=p.category,
            disease=p.disease,
            bucankao=p.bucankao,
            bmi=p.bmi,
            pbf=p.pbf,
            delFlag=p.del_flag,
        ))

    return result


@router.get("/userinfo/avatarZip", tags=["persons"])
def download_avatar_zip(
    session: SessionDep,
    robotCode: str | None = None,
    finalDate: str | None = None,
) -> StreamingResponse:
    """Download avatar zip for robots."""
    query = select(PersonInfo).where(PersonInfo.del_flag == "0")

    if finalDate:
        try:
            final_dt = datetime.strptime(finalDate, "%Y-%m-%d %H:%M:%S")
            final_dt = final_dt.replace(tzinfo=timezone.utc)
            query = select(PersonInfo).where(PersonInfo.updated_at > final_dt)
        except ValueError:
            pass

    persons = session.exec(query).all()

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for p in persons:
            if p.avatar_path and os.path.exists(p.avatar_path):
                ext = os.path.splitext(p.avatar_path)[1]
                zf.write(p.avatar_path, f"{p.card_id}{ext}")

    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/octet-stream",
        headers={"Content-Disposition": "attachment; filename=avatars.zip"},
    )


# ===== Helper =====

def _mark_robots_for_user_update(session: SessionDep) -> None:
    """Mark all robots to send update_user on next heartbeat."""
    robots = session.exec(select(Robot)).all()
    for robot in robots:
        if not robot.pending_command:
            robot.pending_command = "update_user"
            robot.needs_user_update = True
            session.add(robot)
    session.commit()
